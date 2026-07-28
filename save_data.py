#introduciamo anche i dati del pdf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re


# =====================================================================
# COLORI DEI LOTTI
# Un bando multi-lotto occupa molte righe consecutive (esito-191 ne genera
# 96, Esito-205 ne ha 54 divise in 9 lotti): senza un segno visivo i blocchi
# si fondono, e bandi multi-lotto CONSECUTIVI diventano indistinguibili.
#
# Lo schema e' a due livelli:
#   - la FAMIGLIA di colore identifica il BANDO, e cambia da un bando
#     multi-lotto al successivo;
#   - la TONALITA' dentro la famiglia identifica il LOTTO, scurendosi dal
#     primo all'ultimo.
# Cosi' si legge a colpo d'occhio "stesso bando, lotto diverso" senza
# confondere due gare vicine.
#
# Le scale restano CHIARE anche nella tonalita' piu' scura: il testo deve
# restare leggibile e la riga del vincitore (gialla) deve risaltare sopra
# qualunque di questi sfondi.
# I bandi mono-lotto restano su sfondo bianco.
# =====================================================================
FAMIGLIE_COLORE = [
    ["EDF4FC", "DCE9F7", "CBDFF2", "BAD4ED", "A9C9E8", "98BFE3"],  # azzurro
    ["EAF6EF", "D6EDDF", "C2E4CF", "AEDBBF", "9AD2AF", "86C99F"],  # verde
    ["FDF4E6", "FAE9CD", "F7DEB4", "F4D39B", "F1C882", "EEBD69"],  # sabbia
    ["F4EEFA", "E9DDF5", "DECCF0", "D3BBEB", "C8AAE6", "BD99E1"],  # lilla
    ["E8F5F5", "D1EBEB", "BAE1E1", "A3D7D7", "8CCDCD", "75C3C3"],  # acqua
    ["FCEDED", "F9DBDB", "F6C9C9", "F3B7B7", "F0A5A5", "ED9393"],  # rosa
]

GIALLO_VINCITORE = "FFF2A8"
TESTO_NESSUN_INVITATO = "lista invitati non presente"
# Il PDF di esito NON esiste per quella gara: e' un caso diverso dal PDF che
# c'e' ma non elenca gli invitati (TESTO_NESSUN_INVITATO). Distinguerli conta:
# nel primo caso non si sa nulla, nel secondo si sa che la stazione appaltante
# non ha pubblicato la lista.
TESTO_PDF_ASSENTE = "PDF non presente"


def _norm_codice(codice):
    """
    Ripulisce P.IVA / codice fiscale per il confronto: via spazi, punti,
    trattini e prefisso nazionale "IT", tutto maiuscolo.
    """
    if not codice or codice in ("Non presente", "Non trovato"):
        return ""
    pulito = re.sub(r'[\s.\-/]', '', str(codice)).upper()
    if pulito.startswith('IT') and len(pulito) > 2:
        pulito = pulito[2:]
    return pulito


def _codici_aggiudicatario(dati_anac, lotto):
    """
    Codici dell'aggiudicatario, con ANAC che ha la PRECEDENZA e il PDF che fa
    da ripiego quando ANAC non fornisce il dato.

    ANAC pubblica solo il CODICE_FISCALE (mai la P.IVA) e negli RTI/ATI ne
    elenca uno per ogni membro del raggruppamento: si restituisce quindi un
    INSIEME di codici, e l'invitato risulta vincitore se coincide con almeno
    uno di essi — altrimenti si perderebbero le aggiudicazioni ai gruppi.
    """
    codici = set()
    for c in str((dati_anac or {}).get("aggiudicatario_cf", "")).split(","):
        n = _norm_codice(c)
        if n:
            codici.add(n)
    if codici:
        return codici
    # Ripiego sul PDF: nell'archivio 215 aggiudicatari su 226 hanno la P.IVA
    # dichiarata nel documento, quindi il confronto resta possibile quasi
    # sempre anche senza ANAC.
    for campo in ("aggiudicatario_piva", "aggiudicatario_cf"):
        for c in str((lotto or {}).get(campo, "")).split(","):
            n = _norm_codice(c)
            if n:
                codici.add(n)
    return codici


def _nome_aggiudicatario(dati_anac, lotto):
    """
    Nome dell'aggiudicatario: ANAC prima, PDF come ripiego, "Deserto" sui
    lotti andati deserti (nessuna delle due fonti ha un aggiudicatario
    perche' non esiste).
    """
    nome = (dati_anac or {}).get("aggiudicatario", "")
    if nome and nome != "Non presente":
        return nome
    if (lotto or {}).get("deserto"):
        return "Deserto"
    nome = (lotto or {}).get("aggiudicatario_pdf", "")
    return "" if nome in ("Non presente", None) else nome


def _codice_aggiudicatario_visibile(dati_anac, lotto):
    """Codice da mostrare in colonna, con la stessa precedenza ANAC -> PDF."""
    cf = (dati_anac or {}).get("aggiudicatario_cf", "")
    if cf and cf != "Non presente":
        return cf
    if (lotto or {}).get("deserto"):
        return ""
    for campo in ("aggiudicatario_piva", "aggiudicatario_cf"):
        v = (lotto or {}).get(campo, "")
        if v and v != "Non presente":
            return v
    return ""


def _invitati_del_lotto(dati_pdf, lotto):
    """
    Invitati che competono a un lotto: quelli propri del lotto se ci sono,
    altrimenti quelli dichiarati a livello di gara (nei bandi mono-lotto e in
    diversi multi-lotto la lista e' unica per tutta la gara).
    """
    propri = (lotto or {}).get("invitati") or []
    if propri:
        return propri
    return (dati_pdf or {}).get("operatori_invitati") or []


def _etichetta_lotto(lotto, indice, totale):
    """
    "Lotto singolo" oppure "Multilotto - LOTTO N". Il nome del lotto si
    prende dal PDF quando c'e' (i lotti possono essere numerati con salti,
    es. 1-3-4, o essere lettere), altrimenti si ripiega sulla posizione.
    """
    if totale <= 1:
        return "Lotto singolo"
    nome = (lotto or {}).get("nome_lotto")
    return f"Multilotto - {nome}" if nome else f"Multilotto - Lotto {indice}"


def _righe_da_bando(bando, piva_cercata=None):
    """
    Trasforma un bando nelle sue righe di tabella: UNA RIGA PER INVITATO, per
    ogni lotto. E' la forma che permette di filtrare per operatore e contarne
    le ricorrenze, cosa impossibile se gli invitati stessero tutti in una cella.

    Con la ricerca per operatore attiva si emette una sola riga per lotto —
    quella dell'operatore cercato — invece dell'intera lista.
    """
    dati_pagina = bando.get("provincia", {}) or {}
    dati_anac = bando.get("anac", {}) or {}
    dati_pdf = bando.get("pdf") or {}
    lotti = dati_pdf.get("lotti") or [{}]
    # Il ciclo di main.py itera PER CIG e restringe i dati al singolo lotto:
    # la lista qui contiene un elemento solo anche nei multi-lotto. Il numero
    # REALE di lotti arriva in "_totale_lotti", altrimenti si userebbe len()
    # e ogni lotto risulterebbe "Lotto singolo".
    totale_lotti = dati_pdf.get("_totale_lotti") or len(lotti)
    # Gara priva del PDF di esito: non c'e' alcun dato documentale da riportare.
    senza_pdf = not dati_pdf
    cercata = _norm_codice(piva_cercata)

    righe = []
    for i, lotto in enumerate(lotti, 1):
        # CIG del lotto: non esiste un CIG di gara che comprenda tutti i lotti,
        # ogni lotto ha il suo. Si ripiega sul CIG corrente del ciclo quando il
        # PDF non lo dichiara.
        cig = (lotto or {}).get("cig_lotto", "Non presente")
        if not cig or cig == "Non presente":
            cig = bando.get("cig_corrente", "Non trovato")

        # Numero REALE del lotto: quando main.py ha ristretto i dati a un solo
        # lotto, la posizione nella lista e' sempre 1 e tutti i lotti
        # prenderebbero lo stesso colore. Si ricava allora dal nome, che puo'
        # essere numerico ("LOTTO 3") oppure una LETTERA ("LOTTO A", "LOTTO B"):
        # senza il caso delle lettere i bandi a lotti alfabetici finivano tutti
        # sulla prima tonalita'.
        _nome_lotto = str((lotto or {}).get("nome_lotto") or "")
        _m_num = re.search(r'(\d+)', _nome_lotto)
        if _m_num:
            indice_reale = int(_m_num.group(1))
        else:
            _m_let = re.search(r'\b([A-Z])\b\s*$', _nome_lotto.strip().upper())
            indice_reale = (ord(_m_let.group(1)) - ord('A') + 1) if _m_let else i

        codici_vinc = _codici_aggiudicatario(dati_anac, lotto)
        invitati = _invitati_del_lotto(dati_pdf, lotto)

        if cercata:
            invitati = [op for op in invitati
                        if isinstance(op, dict)
                        and cercata in (_norm_codice(op.get("piva")),
                                        _norm_codice(op.get("cf")))]
            if not invitati:
                continue

        if not invitati:
            invitati = [None]  # bando senza lista invitati: una riga comunque

        righe_lotto = []
        for op in invitati:
            if not isinstance(op, dict):
                # Nessun PDF per questa gara: le quattro colonne che ne
                # dipendono lo dicono esplicitamente, invece di restare vuote e
                # confondersi con un PDF che semplicemente non elenca nessuno.
                # Aggiudicatario e codice restano quelli di ANAC, che non
                # dipendono dal documento.
                if senza_pdf:
                    etichetta = TESTO_PDF_ASSENTE
                    nome_inv = piva_inv = cf_inv = TESTO_PDF_ASSENTE
                else:
                    # Il PDF c'e' ma non elenca invitati: si dichiara l'assenza
                    # della lista, e i codici restano VUOTI perche' non c'e'
                    # alcun operatore a cui riferirli.
                    etichetta = _etichetta_lotto(lotto, i, totale_lotti)
                    nome_inv, piva_inv, cf_inv = TESTO_NESSUN_INVITATO, "", ""
                vincitore = False
            else:
                etichetta = _etichetta_lotto(lotto, i, totale_lotti)
                nome_inv = op.get("nome", "")
                piva_inv = "" if op.get("piva") == "Non presente" else (op.get("piva") or "")
                cf_inv = "" if op.get("cf") == "Non presente" else (op.get("cf") or "")
                codici_op = {_norm_codice(piva_inv), _norm_codice(cf_inv)} - {""}
                # Vincitore solo se il confronto e' DAVVERO possibile: senza
                # codici da una delle due parti non si afferma nulla.
                vincitore = bool(codici_op & codici_vinc)
                if vincitore:
                    nome_inv = f"{nome_inv} (VINCITORE)"

            righe_lotto.append({
                "valori": [
                    cig,
                    dati_anac.get("oggetto_gara", ""),
                    dati_pagina.get("tipologia", ""),
                    dati_pagina.get("scelta_contraente", ""),
                    dati_pagina.get("enti", ""),
                    dati_pagina.get("data_pubblicazione", ""),
                    dati_pagina.get("scadenza_manifestazione", ""),
                    dati_pagina.get("data_scadenza", ""),
                    dati_anac.get("cup", ""),
                    dati_anac.get("cod_cpv", ""),
                    dati_anac.get("descrizione_cpv", ""),
                    dati_anac.get("tipo_scelta_contraente", ""),
                    etichetta,
                    nome_inv,
                    piva_inv,
                    cf_inv,
                    _nome_aggiudicatario(dati_anac, lotto),
                    _codice_aggiudicatario_visibile(dati_anac, lotto),
                    dati_anac.get("numero_gara", ""),
                    dati_pagina.get("url_provincia", ""),
                ],
                "vincitore": vincitore,
                "indice_lotto": indice_reale,
                # Chiave del BANDO, per raggruppare le righe da colorare nella
                # stessa famiglia. Il NUMERO_GARA di ANAC identifica la
                # PROCEDURA (i lotti di una stessa gara lo condividono, mentre
                # il CIG e' proprio di ciascuno), ma arriva da ANAC e manca
                # quando la chiamata fallisce o il CIG non e' disponibile:
                # l'URL della pagina, sempre presente, fa da garanzia.
                # Chiave del BANDO, per raggruppare le righe da colorare nella
                # stessa famiglia. L'URL della pagina e' l'unico dato SEMPRE
                # presente e sempre identico per tutti i lotti di una gara.
                # Il NUMERO_GARA di ANAC sembrerebbe piu' appropriato, ma non e'
                # affidabile: di solito i lotti lo condividono, pero' capita che
                # ANAC ne assegni uno DIVERSO a ciascuno (Chiesina Uzzanese, 8
                # lotti registrati come 8 procedure separate), e in quel caso il
                # bando si spezzerebbe in 8 famiglie di colore. Resta come
                # ripiego per l'eventualita' che l'URL manchi.
                "chiave_bando": (dati_pagina.get("url_provincia", "")
                                 or str(dati_anac.get("numero_gara", "")).strip()),
                "multi": totale_lotti > 1,
                "ultima_del_lotto": False,
                "ultima_del_bando": False,
            })

        if righe_lotto:
            righe_lotto[-1]["ultima_del_lotto"] = True
            righe.extend(righe_lotto)

    if righe:
        righe[-1]["ultima_del_bando"] = True
    return righe


def salva_in_excel(lista_bandi, nome_file=None, piva_invitato=None):
    """
    Scrive la tabella dei bandi, con UNA RIGA PER INVITATO.

    piva_invitato, se valorizzato, restringe la tabella: solo i bandi in cui
    l'operatore compare e, per ciascun lotto, la sola riga che lo riguarda.
    """
    if not nome_file:
        ora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_file = f"bandi_pistoia_{ora}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Bandi di Gara"

    intestazioni = [
        "CIG", "Oggetto Gara", "Tipologia", "Scelta Contraente", "Enti",
        "Data Pubblicazione", "Scadenza Manif. Interesse", "Data Scadenza",
        "CUP", "CPV", "Descrizione CPV", "Tipo Scelta Contraente (ANAC)",
        "Tipologia lotto", "Invitato", "P.IVA invitato", "C.F. invitato",
        "Aggiudicatario", "CF/P.IVA Aggiudicatario", "Numero Gara", "URL Bando"
    ]

    stile_intestazione = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    stile_sfondo = PatternFill("solid", start_color="1F4E79")
    stile_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bordo_sottile = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    # Bordo MEDIO sotto l'ultimo invitato di un lotto, SPESSO sotto l'ultima
    # riga del bando: sono i due livelli di raggruppamento della tabella.
    bordo_fine_lotto = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='medium'))
    bordo_fine_bando = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thick'))

    for col, intestazione in enumerate(intestazioni, 1):
        cella = ws.cell(row=1, column=col, value=intestazione)
        cella.font = stile_intestazione
        cella.fill = stile_sfondo
        cella.alignment = stile_centro
        cella.border = bordo_sottile

    font_dati = Font(name="Arial", size=10)
    font_vincitore = Font(name="Arial", size=10, bold=True)
    riempimento_giallo = PatternFill("solid", start_color=GIALLO_VINCITORE)
    allineamento_dati = Alignment(vertical="center", wrap_text=True)

    righe = []
    for bando in lista_bandi:
        righe.extend(_righe_da_bando(bando, piva_invitato))

    # Una FAMIGLIA di colore per ogni bando multi-lotto, assegnata nell'ordine
    # in cui i bandi compaiono: bandi consecutivi ricevono famiglie diverse,
    # cosi' non si confondono fra loro. I mono-lotto non entrano nel giro.
    famiglia_per_bando = {}
    for riga in righe:
        if riga["multi"] and riga["chiave_bando"] not in famiglia_per_bando:
            famiglia_per_bando[riga["chiave_bando"]] = len(famiglia_per_bando) % len(FAMIGLIE_COLORE)

    for riga_idx, riga in enumerate(righe, 2):
        # Mono-lotto su sfondo bianco. Nei multi-lotto il colore dice due cose
        # insieme: la famiglia identifica il BANDO, la tonalita' (via via piu'
        # scura) identifica il LOTTO dentro quel bando.
        if riga["multi"]:
            _scala = FAMIGLIE_COLORE[famiglia_per_bando[riga["chiave_bando"]]]
            colore = _scala[min(riga["indice_lotto"] - 1, len(_scala) - 1)]
        else:
            colore = "FFFFFF"
        riempimento = PatternFill("solid", start_color=colore)
        if riga["ultima_del_bando"]:
            bordo = bordo_fine_bando
        elif riga["ultima_del_lotto"]:
            bordo = bordo_fine_lotto
        else:
            bordo = bordo_sottile

        for col_idx, valore in enumerate(riga["valori"], 1):
            cella = ws.cell(row=riga_idx, column=col_idx, value=valore)
            cella.alignment = allineamento_dati
            cella.border = bordo
            # La riga del vincitore e' in grassetto e su giallo: resta
            # riconoscibile sopra qualunque tonalita' di lotto.
            if riga["vincitore"]:
                cella.font = font_vincitore
                cella.fill = riempimento_giallo
            else:
                cella.font = font_dati
                cella.fill = riempimento

    for r in range(2, len(righe) + 2):
        ws.row_dimensions[r].height = 45

    for col in ws.columns:
        larghezza_max = 0
        for cella in col:
            if cella.value:
                larghezza_max = max(larghezza_max, len(str(cella.value)) * 1.2)
        ws.column_dimensions[col[0].column_letter].width = min(max(larghezza_max, 12), 80)

    ws.column_dimensions['A'].width = 18   # CIG
    ws.column_dimensions['B'].width = 70   # Oggetto Gara
    ws.column_dimensions['N'].width = 45   # Invitato

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(nome_file)
    print(f"\n[+] File Excel salvato: {nome_file}  ({len(righe)} righe da {len(lista_bandi)} bandi)")
    return nome_file